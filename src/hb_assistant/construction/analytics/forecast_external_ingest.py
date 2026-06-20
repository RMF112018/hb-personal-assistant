"""External-forecast file ingestion (Implementation Phase 4).

Accepts an untrusted operator-supplied forecast file as base64 bytes (no ``python-multipart``
dependency; the routes are sync), computes a SHA-256 fingerprint, parses the tabular content with
``openpyxl`` (.xlsx) or the stdlib ``csv`` reader (.csv), and persists an immutable import record
under an isolated, fail-closed **eval-root**. Nothing is auto-projected to the live DB.

Untrusted-file posture: bounded byte/row/column caps, ``data_only`` workbook reads (cell values,
never formulas), no ``eval``/exec, no network. Only the file basename is retained (never a path).
"""

from __future__ import annotations

import base64
import binascii
import csv
import hashlib
import io
import json
import os
import uuid
from datetime import datetime
from pathlib import Path
from typing import Any

from hb_assistant.construction.analytics.forecast_external_dto import ImportPreviewDTO, eval_label

ENV_EVAL_ROOT = "HB_FORECAST_EVAL_ROOT"
ENV_DATA_ROOT = "HB_FORECAST_DATA_ROOT"  # the live/source data root (Phase 3); eval-root must be outside it
ENV_DB_PATH = "HB_FORECAST_DB_PATH"  # read-only v59 source-domain DB for baselines

MAX_UPLOAD_BYTES = 10 * 1024 * 1024
MAX_ROWS = 5000
MAX_COLS = 100
SAMPLE_ROWS = 10
SUPPORTED_SOURCE_SYSTEMS = ("excel", "procore", "sage", "manual", "other")

_IMPORTS_DIRNAME = "imports"
_IMPORT_RECORD = "import_record.json"
_PARSED_ROWS = "parsed_rows.jsonl"
_SOURCE_PREFIX = "source"


class ForecastExternalError(RuntimeError):
    """Raised when external-forecast ingestion/evaluation is misconfigured or input is invalid."""


def resolve_eval_root(override: str | None = None) -> Path:
    """Fail-closed eval-root: absolute, creatable, and never under the live data root / live DB dir."""
    raw = override or os.environ.get(ENV_EVAL_ROOT)
    if not raw:
        raise ForecastExternalError("forecast eval root is not configured")
    p = Path(raw)
    if not p.is_absolute():
        raise ForecastExternalError("forecast eval root must be an absolute path")
    # Defense-in-depth: refuse if the eval-root sits under the live forecast DATA root (where the
    # live packages live). The baseline DB is only ever opened read-only, so its location is not a
    # write hazard and is not guarded here.
    live = os.environ.get(ENV_DATA_ROOT)
    if live and _is_under(p, Path(live)):
        raise ForecastExternalError("forecast eval root must not be under the live data root")
    try:
        p.mkdir(parents=True, exist_ok=True)
    except OSError as exc:
        raise ForecastExternalError("forecast eval root could not be created") from exc
    return p


def _is_under(child: Path, parent: Path) -> bool:
    try:
        c = child.resolve(strict=False)
        r = parent.resolve(strict=False)
        return c == r or c.is_relative_to(r)
    except OSError:
        return True  # fail closed


def _safe_basename(filename: str) -> str:
    base = os.path.basename(str(filename or "").strip()) or "upload"
    # Strip any residual separators / parent refs.
    return base.replace("/", "_").replace("\\", "_").replace("..", "_")


def _detect_format(basename: str) -> str:
    lower = basename.lower()
    if lower.endswith(".xlsx"):
        return "xlsx"
    if lower.endswith(".csv"):
        return "csv"
    raise ForecastExternalError("unsupported file type (expected .xlsx or .csv)")


def _parse_csv(data: bytes) -> tuple[list[str], list[dict[str, Any]]]:
    text = data.decode("utf-8-sig", errors="replace")
    reader = csv.reader(io.StringIO(text))
    rows = list(reader)
    if not rows:
        return [], []
    header = [str(h).strip() for h in rows[0][:MAX_COLS]]
    out: list[dict[str, Any]] = []
    for raw in rows[1 : 1 + MAX_ROWS]:
        cells = raw[:MAX_COLS]
        out.append({header[i]: cells[i] for i in range(min(len(header), len(cells)))})
    return header, out


def _parse_xlsx(data: bytes) -> tuple[list[str], list[dict[str, Any]], list[str]]:
    import openpyxl  # local import; only needed for xlsx

    wb = openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    try:
        sheet_names = list(wb.sheetnames)
        ws = wb[sheet_names[0]]
        rows_iter = ws.iter_rows(values_only=True)
        try:
            first = next(rows_iter)
        except StopIteration:
            return [], [], sheet_names
        header = [str(c).strip() if c is not None else "" for c in list(first)[:MAX_COLS]]
        out: list[dict[str, Any]] = []
        for count, raw in enumerate(rows_iter):
            if count >= MAX_ROWS:
                break
            cells = list(raw)[:MAX_COLS]
            out.append(
                {
                    header[i]: ("" if cells[i] is None else cells[i])
                    for i in range(min(len(header), len(cells)))
                }
            )
        return header, out, sheet_names
    finally:
        wb.close()


class ForecastExternalIngestService:
    """Ingests + previews an uploaded external forecast file into the isolated eval-root."""

    def __init__(self, eval_root: str | None = None) -> None:
        self._eval_root_override = eval_root

    def _eval_root(self) -> Path:
        return resolve_eval_root(self._eval_root_override)

    def _imports_root(self) -> Path:
        root = self._eval_root() / _IMPORTS_DIRNAME
        root.mkdir(parents=True, exist_ok=True)
        return root

    def preview(
        self,
        filename: str,
        content_b64: str,
        source_system: str = "excel",
        period: str | None = None,
    ) -> dict[str, Any]:
        source_system = (source_system or "excel").strip().lower()
        if source_system not in SUPPORTED_SOURCE_SYSTEMS:
            raise ForecastExternalError("unsupported source system")
        basename = _safe_basename(filename)
        fmt = _detect_format(basename)
        try:
            data = base64.b64decode(content_b64 or "", validate=True)
        except (binascii.Error, ValueError) as exc:
            raise ForecastExternalError("uploaded content is not valid base64") from exc
        if not data:
            raise ForecastExternalError("uploaded file is empty")
        if len(data) > MAX_UPLOAD_BYTES:
            raise ForecastExternalError("uploaded file exceeds the size limit")

        file_sha256 = hashlib.sha256(data).hexdigest()
        sheet_names: list[str] = []
        if fmt == "xlsx":
            columns, parsed, sheet_names = _parse_xlsx(data)
        else:
            columns, parsed = _parse_csv(data)
        if not columns:
            raise ForecastExternalError("no columns detected in the uploaded file")

        import_id = uuid.uuid4().hex[:12]
        created_stamp = datetime.now().strftime("%Y%m%d_%H%M%S")  # noqa: DTZ005 local stamp
        imports_root = self._imports_root()
        import_dir = imports_root / import_id
        import_dir.mkdir(parents=True, exist_ok=True)
        # Store the untrusted source bytes immutably + the parsed rows for downstream stages.
        (import_dir / f"{_SOURCE_PREFIX}.{fmt}").write_bytes(data)
        with (import_dir / _PARSED_ROWS).open("w", encoding="utf-8") as fh:
            for row in parsed:
                fh.write(json.dumps(row, default=str, sort_keys=True) + "\n")
        record = {
            "import_id": import_id,
            "created_stamp": created_stamp,
            "source_system": source_system,
            "period": period,
            "source_filename": basename,
            "format": fmt,
            "file_sha256": file_sha256,
            "byte_count": len(data),
            "row_count": len(parsed),
            "columns": columns,
            "sheet_names": sheet_names,
        }
        (import_dir / _IMPORT_RECORD).write_text(
            json.dumps(record, indent=2, sort_keys=True), encoding="utf-8"
        )

        dto = ImportPreviewDTO(
            import_id=import_id,
            display_label=eval_label(source_system, period, created_stamp),
            source_system=source_system,
            period=period,
            source_filename=basename,
            file_sha256=file_sha256,
            byte_count=len(data),
            sheet_names=sheet_names,
            columns=columns,
            sample_rows=parsed[:SAMPLE_ROWS],
            row_count=len(parsed),
        )
        return dto.public()

    # -- shared readers used by the mapping / eval services -------------------

    def read_import_record(self, import_id: str) -> dict[str, Any]:
        path = self._imports_root() / import_id / _IMPORT_RECORD
        try:
            obj = json.loads(path.read_text(encoding="utf-8"))
        except (OSError, ValueError) as exc:
            raise ForecastExternalError(f"unknown import_id: {import_id!r}") from exc
        if not isinstance(obj, dict):
            raise ForecastExternalError(f"unknown import_id: {import_id!r}")
        return obj

    def read_parsed_rows(self, import_id: str) -> list[dict[str, Any]]:
        path = self._imports_root() / import_id / _PARSED_ROWS
        rows: list[dict[str, Any]] = []
        try:
            with path.open("r", encoding="utf-8") as fh:
                for line in fh:
                    line = line.strip()
                    if line:
                        rows.append(json.loads(line))
        except (OSError, ValueError) as exc:
            raise ForecastExternalError(f"unknown import_id: {import_id!r}") from exc
        return rows
