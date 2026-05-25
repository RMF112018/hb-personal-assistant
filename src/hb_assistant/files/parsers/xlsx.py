"""XLSX/XLSMParser: bounded cell values extraction (openpyxl). No macros, no formulas eval."""

from __future__ import annotations

from pathlib import Path
from typing import Any, Dict

try:
    from openpyxl import load_workbook  # type: ignore
except ImportError:
    load_workbook = None  # type: ignore


class XLSXParser:
    MAX_ROWS_PER_SHEET = 50
    MAX_COLS = 20

    def parse(self, path: Path, max_chars: int = 8000) -> Dict[str, Any]:
        if load_workbook is None:
            raise RuntimeError("openpyxl not installed")
        try:
            wb = load_workbook(str(path), data_only=True, read_only=True)
            parts: list[str] = []
            total = 0
            sheets = 0
            for sheet_name in wb.sheetnames:
                sheets += 1
                ws = wb[sheet_name]
                parts.append(f"--- {sheet_name} ---")
                row_count = 0
                for row in ws.iter_rows(min_row=1, max_row=self.MAX_ROWS_PER_SHEET, max_col=self.MAX_COLS, values_only=True):
                    vals = [str(v) if v is not None else "" for v in row]
                    line = " | ".join(vals).strip()
                    if line:
                        parts.append(line)
                        total += len(line) + 1
                        if total > max_chars:
                            break
                    row_count += 1
                    if row_count >= self.MAX_ROWS_PER_SHEET:
                        break
                if total > max_chars:
                    break
            excerpt = "\n".join(parts)[:max_chars]
            return {
                "text_excerpt": excerpt,
                "char_count": len(excerpt),
                "sheet_count": sheets,
                "rows_sampled": self.MAX_ROWS_PER_SHEET,
            }
        except Exception as e:
            msg = str(e)[:200]
            fc = "parser_error"
            return {"text_excerpt": "", "char_count": 0, "error": msg, "failure_code": fc}
