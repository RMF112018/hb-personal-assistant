"""Phase B / B2 — subprocess-isolated, bounded parser extraction.

Proves the read-time parser boundary (``files/parsers/isolated``) extracts PDF/DOCX/XLSX/EML in a child
process and that the PARENT SURVIVES a hostile file — timeout, crash-by-signal, nonzero exit, malformed
payload, oversize output, oversize input — always returning a deterministic classification and leaving
no lingering child. Includes a process-group containment proof (a worker that forks a grandchild is
killed as a group on timeout). Also exercises the provider's complete-read path for each real format.
"""

from __future__ import annotations

import hashlib
import multiprocessing
import os
import signal
import sqlite3
import time
from datetime import datetime, timezone
from pathlib import Path

import pytest

from hb_assistant.files.parsers import isolated as iso
from hb_assistant.obsidian_mcp import source_connector_service as svc
from hb_assistant.obsidian_mcp.config import ExternalSourceRoot, ObsidianMcpConfig
from hb_assistant.obsidian_mcp.source_connector_models import encode_source_ref
from hb_assistant.obsidian_mcp.source_index_repository import SourceIndexRepository
from hb_assistant.store.migrator import SQLiteMigrator

_LIMITS = {"max_input_bytes": 25_000_000, "max_output_bytes": 1_000_000, "timeout_s": 15.0,
           "max_memory_mb": 512}

_MINIMAL_TEXT_PDF = b"""%PDF-1.4
1 0 obj<</Type/Catalog/Pages 2 0 R>>endobj
2 0 obj<</Type/Pages/Kids[3 0 R]/Count 1>>endobj
3 0 obj<</Type/Page/Parent 2 0 R/MediaBox[0 0 300 200]/Contents 4 0 R/Resources<</Font<</F1 5 0 R>>>>>>endobj
4 0 obj<</Length 58>>stream
BT /F1 18 Tf 20 120 Td (Pay Application 12 Contract A) Tj ET
endstream endobj
5 0 obj<</Type/Font/Subtype/Type1/BaseFont/Helvetica>>endobj
xref
0 6
0000000000 65535 f
trailer<</Root 1 0 R/Size 6>>
startxref
0
%%EOF"""


# ---------- fixtures generated in-test (no committed binaries) ----------

def _make_pdf(p: Path) -> Path:
    p.write_bytes(_MINIMAL_TEXT_PDF)
    return p


def _make_docx(p: Path) -> Path:
    import docx

    d = docx.Document()
    d.add_paragraph("Pay Application 12 for Contract A")
    d.save(p)
    return p


def _make_xlsx(p: Path) -> Path:
    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active
    ws["A1"] = "Invoice"
    ws["B1"] = 12345
    wb.save(p)
    return p


def _make_eml(p: Path) -> Path:
    p.write_text(
        "From: a@b.com\nTo: c@d.com\nSubject: Pay Application\n"
        "Content-Type: text/plain\n\nBody: pay application 12 total due.\n"
    )
    return p


# ---------- module-level test workers (must be top-level for spawn pickling) ----------

def _worker_sleep(path_str, ext, mob, mmb, timeout_s, send):
    os.setsid()
    time.sleep(300)


def _worker_segfault(path_str, ext, mob, mmb, timeout_s, send):
    os.setsid()
    os.kill(os.getpid(), signal.SIGSEGV)


def _worker_nonzero(path_str, ext, mob, mmb, timeout_s, send):
    os.setsid()
    os._exit(7)


def _worker_malformed(path_str, ext, mob, mmb, timeout_s, send):
    os.setsid()
    send.send("not-a-dict")
    send.close()


def _worker_sigkill(path_str, ext, mob, mmb, timeout_s, send):
    os.setsid()
    os.kill(os.getpid(), signal.SIGKILL)  # unsolicited SIGKILL (NOT our timeout path)


def _worker_fork_grandchild(path_str, ext, mob, mmb, timeout_s, send):
    os.setsid()  # become group leader so a killpg reaches the grandchild too
    pid = os.fork()
    if pid == 0:  # grandchild
        pidfile = os.environ["HB_TEST_GC_PIDFILE"]
        Path(pidfile).write_text(str(os.getpid()))
        time.sleep(300)
        os._exit(0)
    time.sleep(300)  # worker hangs -> parent times out and kills the whole group


def _pid_alive(pid: int) -> bool:
    try:
        os.kill(pid, 0)
        return True
    except ProcessLookupError:
        return False
    except PermissionError:
        return True


# ---------- direct isolation-module tests ----------

@pytest.mark.parametrize("ext,maker,needle", [
    ("pdf", _make_pdf, "Pay Application 12"),
    ("docx", _make_docx, "Pay Application 12"),
    ("xlsx", _make_xlsx, "Invoice"),
    ("eml", _make_eml, "pay application 12"),
])
def test_isolated_parses_real_fixture(tmp_path, ext, maker, needle) -> None:
    p = maker(tmp_path / f"f.{ext}")
    r = iso.extract_for_complete_read(p, ext, **_LIMITS)
    assert r.status == "ok"
    assert needle.lower() in (r.text or "").lower()
    assert not multiprocessing.active_children()


def test_corrupt_pdf_does_not_crash(tmp_path) -> None:
    p = tmp_path / "bad.pdf"
    p.write_bytes(b"%PDF-1.4\n%garbage not a real pdf\n" + os.urandom(64))
    r = iso.extract_for_complete_read(p, "pdf", **_LIMITS)
    assert r.status in ("parser_failed", "parser_output_too_large")
    assert not multiprocessing.active_children()


def test_corrupt_xlsx_does_not_crash(tmp_path) -> None:
    p = tmp_path / "bad.xlsx"
    p.write_bytes(b"PK\x03\x04 not really a zip/xlsx " + os.urandom(64))
    r = iso.extract_for_complete_read(p, "xlsx", **_LIMITS)
    assert r.status == "parser_failed"
    assert not multiprocessing.active_children()


def test_output_over_budget_is_not_ok(tmp_path) -> None:
    p = _make_docx(tmp_path / "big.docx")
    r = iso.extract_for_complete_read(p, "docx", max_input_bytes=25_000_000, max_output_bytes=5,
                                      timeout_s=15.0, max_memory_mb=512)
    assert r.status == "parser_output_too_large"
    assert r.text is None
    assert r.observed_output_bytes_lower_bound and r.observed_output_bytes_lower_bound > 5


def test_input_over_budget_is_too_large_no_child(tmp_path) -> None:
    p = _make_docx(tmp_path / "in.docx")
    before = len(multiprocessing.active_children())
    r = iso.extract_for_complete_read(p, "docx", max_input_bytes=10, max_output_bytes=1_000_000,
                                      timeout_s=15.0, max_memory_mb=512)
    assert r.status == "too_large"
    assert len(multiprocessing.active_children()) == before  # never spawned


def test_timeout_enforced(tmp_path) -> None:
    p = _make_docx(tmp_path / "t.docx")
    t0 = time.monotonic()
    r = iso.extract_for_complete_read(p, "docx", max_input_bytes=25_000_000, max_output_bytes=1_000_000,
                                      timeout_s=1.0, max_memory_mb=512, _worker=_worker_sleep)
    assert r.status == "parser_timeout"
    assert time.monotonic() - t0 < 20  # actually killed, not run to completion
    assert not multiprocessing.active_children()


def test_segfault_classified_failed_parent_survives(tmp_path) -> None:
    p = _make_docx(tmp_path / "t.docx")
    r = iso.extract_for_complete_read(p, "docx", max_input_bytes=25_000_000, max_output_bytes=1_000_000,
                                      timeout_s=5.0, max_memory_mb=512, _worker=_worker_segfault)
    assert r.status == "parser_failed"
    # parent still healthy: a normal extraction works right after
    ok = iso.extract_for_complete_read(p, "docx", **_LIMITS)
    assert ok.status == "ok"


def test_nonzero_exit_classified_failed(tmp_path) -> None:
    p = _make_docx(tmp_path / "t.docx")
    r = iso.extract_for_complete_read(p, "docx", max_input_bytes=25_000_000, max_output_bytes=1_000_000,
                                      timeout_s=5.0, max_memory_mb=512, _worker=_worker_nonzero)
    assert r.status == "parser_failed"


def test_malformed_payload_classified_failed(tmp_path) -> None:
    p = _make_docx(tmp_path / "t.docx")
    r = iso.extract_for_complete_read(p, "docx", max_input_bytes=25_000_000, max_output_bytes=1_000_000,
                                      timeout_s=5.0, max_memory_mb=512, _worker=_worker_malformed)
    assert r.status == "parser_failed"


def test_process_group_kills_grandchild(tmp_path, monkeypatch) -> None:
    pidfile = tmp_path / "gc.pid"
    monkeypatch.setenv("HB_TEST_GC_PIDFILE", str(pidfile))
    p = _make_docx(tmp_path / "t.docx")
    r = iso.extract_for_complete_read(p, "docx", max_input_bytes=25_000_000, max_output_bytes=1_000_000,
                                      timeout_s=1.5, max_memory_mb=512, _worker=_worker_fork_grandchild)
    assert r.status == "parser_timeout"
    # the forked grandchild must also be dead (process-group kill), not orphaned
    assert pidfile.exists(), "grandchild never recorded its pid"
    gc_pid = int(pidfile.read_text().strip())
    deadline = time.monotonic() + 3.0
    while _pid_alive(gc_pid) and time.monotonic() < deadline:
        time.sleep(0.05)
    assert not _pid_alive(gc_pid), "grandchild survived the process-group kill"


def test_unsupported_ext_no_spawn(tmp_path) -> None:
    p = tmp_path / "x.xer"
    p.write_text("dummy")
    r = iso.extract_for_complete_read(p, "xer", **_LIMITS)
    assert r.status == "unsupported_format"


# ---------- PB-008: CPU rlimit derived from the wall timeout ----------

def test_cpu_limits_derived_from_timeout() -> None:
    # The soft CPU limit tracks timeout_s (sits a couple seconds above it), not a fixed 60/90 constant.
    assert iso._cpu_limits(20.0) == (22, 25)
    assert iso._cpu_limits(5.0) == (7, 10)
    soft, hard = iso._cpu_limits(0.0)  # floored, still valid + ordered
    assert soft >= 1 and hard > soft


# ---------- PB-009: SIGKILL is ambiguous -> parser_failed; SIGXCPU -> resource_exceeded ----------

def test_sigkill_classified_parser_failed(tmp_path) -> None:
    # An UNSOLICITED SIGKILL (OOM killer / operator kill — not our timeout path) is not provably resource
    # exhaustion, so it classifies as parser_failed, never parser_resource_exceeded.
    p = _make_docx(tmp_path / "t.docx")
    r = iso.extract_for_complete_read(p, "docx", max_input_bytes=25_000_000, max_output_bytes=1_000_000,
                                      timeout_s=5.0, max_memory_mb=512, _worker=_worker_sigkill)
    assert r.status == "parser_failed"
    assert r.failure_code == f"signal_{int(signal.SIGKILL)}"


def test_sigxcpu_classified_resource_exceeded() -> None:
    # The CPU rlimit backstop (SIGXCPU) IS provable resource exhaustion.
    r = iso._classify_dead_child(-int(signal.SIGXCPU))
    assert r.status == "parser_resource_exceeded"
    assert r.failure_code == f"signal_{int(signal.SIGXCPU)}"


# ---------- provider-level complete reads of each real format ----------

def _index_and_trust(db: str, config: ObsidianMcpConfig, root_key: str, rel_path: str,
                     abs_file: Path, ext: str) -> str:
    from hb_assistant.obsidian_mcp.source_indexer import _root_fingerprint

    st = abs_file.stat()
    now = datetime.now(timezone.utc).replace(microsecond=0).isoformat()
    cfg_root = next(r for r in config.external_sources if r.source_root_key == root_key)
    fp = _root_fingerprint(cfg_root, config)
    rph = hashlib.sha256(str(Path(cfg_root.path)).encode()).hexdigest()[:32]
    sid = SourceIndexRepository(db).upsert_source_file({
        "source_kind": "external_file",
        "source_root_key": root_key,
        "rel_path": rel_path,
        "file_ext": ext,
        "size_bytes": st.st_size,
        "mtime_ns": st.st_mtime_ns,
        "content_sha256": "d",
        "extraction_status": "pending",
        "extraction_disposition": "metadata_only",
    })
    with sqlite3.connect(db) as c:
        c.execute("INSERT OR REPLACE INTO source_index_scan_generations(generation_id, root_key, status, "
                  "root_path_hash, policy_fingerprint, started_at, updated_at, metadata_walk_completed_at, "
                  "reconciliation_completed_at, finished_at) VALUES(?,?,?,?,?,?,?,?,?,?)",
                  (f"gen-{root_key}", root_key, "completed", rph, fp, now, now, now, now, now))
        c.commit()
    return sid


@pytest.mark.parametrize("ext,maker,needle", [
    ("pdf", _make_pdf, "Pay Application 12"),
    ("docx", _make_docx, "Pay Application 12"),
    ("xlsx", _make_xlsx, "Invoice"),
    ("eml", _make_eml, "pay application 12"),
])
def test_provider_complete_read_parser_formats(tmp_path, ext, maker, needle) -> None:
    db = str(tmp_path / "db.sqlite")
    SQLiteMigrator(db_path=db).apply()
    work = tmp_path / "work"
    work.mkdir()
    f = maker(work / f"doc.{ext}")
    config = ObsidianMcpConfig(external_sources=[ExternalSourceRoot(source_root_key="work", path=str(work))])
    sid = _index_and_trust(db, config, "work", f"doc.{ext}", f, ext)
    r = svc.read_source_file(SourceIndexRepository(db), config, source_ref=encode_source_ref(sid),
                             mode="complete")
    assert r["retrieval_state"] == "complete"
    assert r["content_state"] == "extracted_content"
    assert needle.lower() in (r["content"] or "").lower()
    assert str(tmp_path) not in str(r)
